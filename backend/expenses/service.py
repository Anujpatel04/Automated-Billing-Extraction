from extensions.mongodb import mongodb
from expenses.models import ExpenseModel, ExpenseStatus
from ai.bill_extractor import BillExtractor
from storage.file_manager import FileManager
from utils.responses import success_response, error_response
from bson import ObjectId
from datetime import datetime
from typing import Optional, Dict, Any
from flask import send_file
import logging
import os
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class ExpenseService:
    @staticmethod
    def create_expense(user_id: str, file) -> tuple:
        try:
            is_valid, error_msg = FileManager.validate_file(file)
            if not is_valid:
                return error_response(error_msg, 400)
            
            image_path = FileManager.save_file(file, user_id)
            if not image_path:
                return error_response("Failed to save file", 500)
            
            try:
                extracted_data = BillExtractor.extract_bill_data(image_path)
            except Exception as e:
                logger.error(f"Bill extraction error: {str(e)}", exc_info=True)
                FileManager.delete_file(image_path)
                return error_response(
                    f"Failed to extract bill data: {str(e)}. Please ensure the image is clear and contains a valid bill.",
                    400
                )
            
            if not extracted_data:
                FileManager.delete_file(image_path)
                logger.warning(f"Extraction returned empty data for {image_path}")
                return error_response(
                    "Failed to extract bill data from image. Please ensure the image is clear and contains readable text.",
                    400
                )
            
            is_valid, error_msg = ExpenseModel.validate_extracted_data(extracted_data)
            if not is_valid:
                FileManager.delete_file(image_path)
                logger.warning(f"Validation failed for {image_path}: {error_msg}")
                return error_response(
                    f"Invalid extracted data: {error_msg}. Please ensure the bill image is clear and contains visible date and amount information.",
                    400
                )
            
            expense_doc = ExpenseModel.create_expense(
                user_id=user_id,
                image_path=image_path,
                extracted_data=extracted_data,
                status=ExpenseStatus.PENDING
            )
            
            expenses_collection = mongodb.get_collection('expenses')
            result = expenses_collection.insert_one(expense_doc)
            expense_id = str(result.inserted_id)
            
            expense = expenses_collection.find_one({'_id': result.inserted_id})
            
            logger.info(f"Expense created: {expense_id} for user: {user_id}")
            
            return success_response(
                "Expense uploaded successfully",
                ExpenseModel.format_expense_response(expense),
                201
            )
            
        except Exception as e:
            logger.error(f"Error creating expense: {str(e)}")
            return error_response("Failed to create expense", 500)
    
    @staticmethod
    def get_user_expenses(user_id: str, status: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None) -> tuple:
        try:
            expenses_collection = mongodb.get_collection('expenses')
            
            query = {'user_id': ObjectId(user_id)}
            if status:
                query['status'] = status
            
            if date_from or date_to:
                query['created_at'] = {}
                if date_from:
                    try:
                        date_from_obj = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                        query['created_at']['$gte'] = date_from_obj
                    except ValueError:
                        pass
                if date_to:
                    try:
                        date_to_obj = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                        query['created_at']['$lte'] = date_to_obj
                    except ValueError:
                        pass
            
            expenses = list(expenses_collection.find(query).sort('created_at', -1))
            
            formatted_expenses = [
                ExpenseModel.format_expense_response(exp) for exp in expenses
            ]
            
            return success_response(
                "Expenses retrieved successfully",
                {'expenses': formatted_expenses, 'count': len(formatted_expenses)}
            )
            
        except Exception as e:
            logger.error(f"Error getting user expenses: {str(e)}")
            return error_response("Failed to retrieve expenses", 500)
    
    @staticmethod
    def get_all_expenses(
        user_id: Optional[str] = None,
        status: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None
    ) -> tuple:
        try:
            expenses_collection = mongodb.get_collection('expenses')
            
            query = {}
            
            if user_id:
                query['user_id'] = ObjectId(user_id)
            
            if status:
                query['status'] = status
            
            if date_from or date_to:
                query['created_at'] = {}
                if date_from:
                    query['created_at']['$gte'] = date_from
                if date_to:
                    query['created_at']['$lte'] = date_to
            
            expenses = list(expenses_collection.find(query).sort('created_at', -1))
            
            users_collection = mongodb.get_collection('users')
            for expense in expenses:
                user = users_collection.find_one({'_id': expense['user_id']})
                expense['user_email'] = user['email'] if user else 'Unknown'
            
            formatted_expenses = [
                ExpenseModel.format_expense_response(exp) for exp in expenses
            ]
            
            return success_response(
                "Expenses retrieved successfully",
                {'expenses': formatted_expenses, 'count': len(formatted_expenses)}
            )
            
        except Exception as e:
            logger.error(f"Error getting all expenses: {str(e)}")
            return error_response("Failed to retrieve expenses", 500)
    
    @staticmethod
    def update_expense_status(expense_id: str, status: str, notes: Optional[str] = None) -> tuple:
        try:
            if status not in [ExpenseStatus.APPROVED, ExpenseStatus.REJECTED, ExpenseStatus.PENDING]:
                return error_response("Invalid status. Must be 'approved', 'rejected', or 'pending'", 400)
            
            expenses_collection = mongodb.get_collection('expenses')
            
            update_data = {
                'status': status,
                'updated_at': datetime.utcnow()
            }
            
            if notes is not None:
                update_data['hr_notes'] = notes.strip() if notes else None
            
            result = expenses_collection.update_one(
                {'_id': ObjectId(expense_id)},
                {'$set': update_data}
            )
            
            if result.matched_count == 0:
                return error_response("Expense not found", 404)
            
            expense = expenses_collection.find_one({'_id': ObjectId(expense_id)})
            
            logger.info(f"Expense status updated: {expense_id} to {status} with notes: {bool(notes)}")
            
            status_messages = {
                'approved': 'approved',
                'rejected': 'rejected',
                'pending': 'set to pending'
            }
            message = f"Expense {status_messages.get(status, status)} successfully"
            
            return success_response(
                message,
                ExpenseModel.format_expense_response(expense)
            )
            
        except Exception as e:
            logger.error(f"Error updating expense status: {str(e)}")
            return error_response("Failed to update expense status", 500)
    
    @staticmethod
    def export_expenses(user_id: str, format_type: str, status: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None) -> tuple:
        try:
            expenses_collection = mongodb.get_collection('expenses')
            
            query = {'user_id': ObjectId(user_id)}
            if status:
                query['status'] = status
            
            if date_from or date_to:
                query['created_at'] = {}
                if date_from:
                    try:
                        date_from_obj = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                        query['created_at']['$gte'] = date_from_obj
                    except ValueError:
                        pass
                if date_to:
                    try:
                        date_to_obj = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                        query['created_at']['$lte'] = date_to_obj
                    except ValueError:
                        pass
            
            expenses = list(expenses_collection.find(query).sort('created_at', -1))
            
            if not expenses:
                return error_response("No expenses found to export", 404)
            
            # Prepare data for export
            export_data = []
            for exp in expenses:
                extracted = exp.get('extracted_data', {})
                amount = extracted.get('Bill Amount (INR)') or extracted.get('Bill Amount') or '0'
                amount_clean = str(amount).replace('₹', '').replace(',', '').replace('$', '').replace('€', '').strip()
                
                export_data.append({
                    'Date': extracted.get('Date', ''),
                    'Vendor': extracted.get('Details', ''),
                    'Bill Type': extracted.get('Bill Type', ''),
                    'Amount (INR)': amount_clean,
                    'Status': exp.get('status', ''),
                    'HR Notes': exp.get('hr_notes', ''),
                    'Created At': exp.get('created_at', '').strftime('%Y-%m-%d %H:%M:%S') if exp.get('created_at') else '',
                    'Updated At': exp.get('updated_at', '').strftime('%Y-%m-%d %H:%M:%S') if exp.get('updated_at') else '',
                })
            
            if format_type.lower() == 'csv':
                df = pd.DataFrame(export_data)
                output = io.StringIO()
                df.to_csv(output, index=False)
                output.seek(0)
                
                from flask import Response
                return Response(
                    output.getvalue(),
                    mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=expenses_{datetime.now().strftime("%Y%m%d")}.csv'}
                )
            else:  # Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Expenses"
                
                # Header row
                headers = list(export_data[0].keys())
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Data rows
                for row_idx, row_data in enumerate(export_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                from flask import Response
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=expenses_{datetime.now().strftime("%Y%m%d")}.xlsx'}
                )
                
        except Exception as e:
            logger.error(f"Error exporting expenses: {str(e)}", exc_info=True)
            return error_response("Failed to export expenses", 500)
    
    @staticmethod
    def download_expense_file(expense_id: str, user_id: str) -> tuple:
        try:
            expenses_collection = mongodb.get_collection('expenses')
            expense = expenses_collection.find_one({'_id': ObjectId(expense_id)})
            
            if not expense:
                return error_response("Expense not found", 404)
            
            # Check if user owns this expense or is HR
            if str(expense['user_id']) != user_id:
                # Check if user is HR
                users_collection = mongodb.get_collection('users')
                user = users_collection.find_one({'_id': ObjectId(user_id)})
                if not user or user.get('role') != 'HR':
                    return error_response("Unauthorized access", 403)
            
            image_path = expense.get('image_path')
            if not image_path or not os.path.exists(image_path):
                return error_response("File not found", 404)
            
            from flask import send_file
            return send_file(
                image_path,
                as_attachment=True,
                download_name=os.path.basename(image_path)
            )
            
        except Exception as e:
            logger.error(f"Error downloading file: {str(e)}")
            return error_response("Failed to download file", 500)
    
    @staticmethod
    def bulk_update_status(expense_ids: list, status: str, notes: Optional[str] = None) -> tuple:
        try:
            if status not in [ExpenseStatus.APPROVED, ExpenseStatus.REJECTED, ExpenseStatus.PENDING]:
                return error_response("Invalid status. Must be 'approved', 'rejected', or 'pending'", 400)
            
            expenses_collection = mongodb.get_collection('expenses')
            
            object_ids = []
            for exp_id in expense_ids:
                try:
                    object_ids.append(ObjectId(exp_id))
                except:
                    continue
            
            if not object_ids:
                return error_response("No valid expense IDs provided", 400)
            
            update_data = {
                'status': status,
                'updated_at': datetime.utcnow()
            }
            
            if notes is not None:
                update_data['hr_notes'] = notes.strip() if notes else None
            
            result = expenses_collection.update_many(
                {'_id': {'$in': object_ids}},
                {'$set': update_data}
            )
            
            logger.info(f"Bulk update: {result.modified_count} expenses updated to {status}")
            
            return success_response(
                f"Successfully updated {result.modified_count} expense(s)",
                {'updated_count': result.modified_count, 'status': status}
            )
            
        except Exception as e:
            logger.error(f"Error bulk updating expenses: {str(e)}")
            return error_response("Failed to bulk update expenses", 500)
    
    @staticmethod
    def export_all_expenses(format_type: str, status: Optional[str] = None, user_id: Optional[str] = None, date_from: Optional[datetime] = None, date_to: Optional[datetime] = None) -> tuple:
        try:
            expenses_collection = mongodb.get_collection('expenses')
            users_collection = mongodb.get_collection('users')
            
            query = {}
            if status:
                query['status'] = status
            if user_id:
                query['user_id'] = ObjectId(user_id)
            if date_from or date_to:
                query['created_at'] = {}
                if date_from:
                    query['created_at']['$gte'] = date_from
                if date_to:
                    query['created_at']['$lte'] = date_to
            
            expenses = list(expenses_collection.find(query).sort('created_at', -1))
            
            if not expenses:
                return error_response("No expenses found to export", 404)
            
            # Get user emails
            user_map = {}
            for expense in expenses:
                user_id_obj = expense['user_id']
                if user_id_obj not in user_map:
                    user = users_collection.find_one({'_id': user_id_obj})
                    user_map[user_id_obj] = user['email'] if user else 'Unknown'
            
            # Prepare data for export
            export_data = []
            for exp in expenses:
                extracted = exp.get('extracted_data', {})
                amount = extracted.get('Bill Amount (INR)') or extracted.get('Bill Amount') or '0'
                amount_clean = str(amount).replace('₹', '').replace(',', '').replace('$', '').replace('€', '').strip()
                
                export_data.append({
                    'User Email': user_map.get(exp['user_id'], 'Unknown'),
                    'Date': extracted.get('Date', ''),
                    'Vendor': extracted.get('Details', ''),
                    'Bill Type': extracted.get('Bill Type', ''),
                    'Amount (INR)': amount_clean,
                    'Status': exp.get('status', ''),
                    'HR Notes': exp.get('hr_notes', ''),
                    'Created At': exp.get('created_at', '').strftime('%Y-%m-%d %H:%M:%S') if exp.get('created_at') else '',
                    'Updated At': exp.get('updated_at', '').strftime('%Y-%m-%d %H:%M:%S') if exp.get('updated_at') else '',
                })
            
            if format_type.lower() == 'csv':
                df = pd.DataFrame(export_data)
                output = io.StringIO()
                df.to_csv(output, index=False)
                output.seek(0)
                
                from flask import Response
                return Response(
                    output.getvalue(),
                    mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=all_expenses_{datetime.now().strftime("%Y%m%d")}.csv'}
                )
            else:  # Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "All Expenses"
                
                # Header row
                headers = list(export_data[0].keys())
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Data rows
                for row_idx, row_data in enumerate(export_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                from flask import Response
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=all_expenses_{datetime.now().strftime("%Y%m%d")}.xlsx'}
                )
                
        except Exception as e:
            logger.error(f"Error exporting all expenses: {str(e)}", exc_info=True)
            return error_response("Failed to export expenses", 500)

